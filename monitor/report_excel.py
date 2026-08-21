# -*- coding: utf-8 -*-
"""生成 Excel 周报（多账户架构，2026-08-20 起按账户分别输出）。

- 默认账户（见 db.DEFAULT_ACCOUNT）：output/素材监控周报.xlsx
- 其他账户：output/素材监控周报-<账户>.xlsx

数据源为天级（daily_stats），此处按 ISO 自然周（周一~周日）聚合展示。

Sheet 清单：
  1. 汇总总览        —— 每周趋势(含环比/新指标)、来源结构、TOP10
  2. 素材生命周期明细 —— 每个素材一张卡：阶段/预警/累计/最新周/环比/新指标
  3. 新增素材-创建口径 —— 按素材创建时间落在该周
  4. 新增素材-首现口径 —— 按首次出现在周报中
  5. 修订记录        —— 官方回调后同天覆盖的痕迹
"""
import os
import sys
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import get_conn, DEFAULT_ACCOUNT
from accounts import discover_accounts, account_has_data
from metrics import (load_all, weekly_trend, source_structure, material_cards,
                     new_week_report, get_week_end_map)

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "output")


def excel_out_path(account):
    """账户 -> 周报文件路径（默认账户沿用旧文件名）。"""
    if account == DEFAULT_ACCOUNT:
        return os.path.join(OUT_DIR, "素材监控周报.xlsx")
    return os.path.join(OUT_DIR, f"素材监控周报-{account}.xlsx")

# ---------- 样式 ----------
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
BODY_FONT = Font(name="微软雅黑", size=10)
SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

NUM_FMT = "#,##0.00"
INT_FMT = "#,##0"
PCT_FMT = "0.00%"


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26


def _header(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def _section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    return row + 1


def _pct(v):
    return v if v is None else round(v, 4)


def _fmt_cell(ws, r, j, v, fmt=None, bold=False, center=False):
    c = ws.cell(row=r, column=j, value=v)
    c.font = Font(name="微软雅黑", size=10, bold=bold)
    if fmt:
        c.number_format = fmt
    if center:
        c.alignment = CENTER
    c.border = BORDER
    return c


def build_overview(ws, weeks, mats, stats, account):
    trend = weekly_trend(stats, weeks)
    _title(ws, f"千川素材监控 · 汇总总览（{account} · 天级聚合为周）", 13)
    r = 3
    _section(ws, r, "一、每周趋势（元；退款率/结算率为按成交金额加权；周 = ISO 自然周）", 13); r += 1
    _header(ws, r, ["周（起始日）", "素材条数", "整体消耗", "整体成交金额", "整体ROI",
                    "净成交金额", "净成交ROI", "净成交订单数", "1小时退款率", "结算率",
                    "消耗环比", "成交金额环比", "整体ROI环比"]); r += 1
    for t in trend:
        _fmt_cell(ws, r, 1, t["week"])
        _fmt_cell(ws, r, 2, t["count"])
        _fmt_cell(ws, r, 3, t["spend"], NUM_FMT)
        _fmt_cell(ws, r, 4, t["pay"], NUM_FMT)
        _fmt_cell(ws, r, 5, t["roi"], "0.00")
        _fmt_cell(ws, r, 6, t["net_pay"], NUM_FMT)
        _fmt_cell(ws, r, 7, t["net_roi"], "0.00")
        _fmt_cell(ws, r, 8, t["order_count"], INT_FMT)
        _fmt_cell(ws, r, 9, _pct(t["refund_rate_1h"]), PCT_FMT)
        _fmt_cell(ws, r, 10, _pct(t["settle_rate"]), PCT_FMT)
        _fmt_cell(ws, r, 11, _pct(t.get("spend_delta")), PCT_FMT)
        _fmt_cell(ws, r, 12, _pct(t.get("pay_delta")), PCT_FMT)
        _fmt_cell(ws, r, 13, _pct(t.get("roi_delta")), PCT_FMT)
        r += 1
    # 累计行
    tot_spend = sum(t["spend"] for t in trend)
    tot_pay = sum(t["pay"] for t in trend)
    tot_net_pay = sum(t["net_pay"] for t in trend)
    tot_order = sum(t["order_count"] for t in trend)
    _fmt_cell(ws, r, 1, "累计", bold=True)
    _fmt_cell(ws, r, 2, sum(t["count"] for t in trend), bold=True)
    _fmt_cell(ws, r, 3, round(tot_spend, 2), NUM_FMT, bold=True)
    _fmt_cell(ws, r, 4, round(tot_pay, 2), NUM_FMT, bold=True)
    _fmt_cell(ws, r, 5, round(tot_pay / tot_spend, 4) if tot_spend else 0, "0.00", bold=True)
    _fmt_cell(ws, r, 6, round(tot_net_pay, 2), NUM_FMT, bold=True)
    _fmt_cell(ws, r, 7, round(tot_net_pay / tot_spend, 4) if tot_spend else 0, "0.00", bold=True)
    _fmt_cell(ws, r, 8, round(tot_order, 2), INT_FMT, bold=True)
    _fmt_cell(ws, r, 9, None)
    _fmt_cell(ws, r, 10, None)
    _fmt_cell(ws, r, 11, None)
    _fmt_cell(ws, r, 12, None)
    _fmt_cell(ws, r, 13, None)
    r += 2

    # 来源结构
    _section(ws, r, "二、素材来源结构（按消耗）", 5); r += 1
    _header(ws, r, ["素材来源", "条数", "消耗", "整体成交金额", "消耗占比"]); r += 1
    for a in source_structure(stats):
        _fmt_cell(ws, r, 1, a["source"])
        _fmt_cell(ws, r, 2, a["count"])
        _fmt_cell(ws, r, 3, a["spend"], NUM_FMT)
        _fmt_cell(ws, r, 4, a["pay"], NUM_FMT)
        _fmt_cell(ws, r, 5, a["ratio"], PCT_FMT)
        r += 1
    r += 1

    # TOP10
    _section(ws, r, "三、累计消耗 TOP10 素材", 8); r += 1
    _header(ws, r, ["素材名称", "素材ID", "来源", "累计消耗", "累计成交金额",
                    "整体ROI", "累计净成交", "净成交ROI"]); r += 1
    cards = material_cards(weeks, mats, stats)
    for card in cards[:10]:
        name = card["name"] + ("（聚合）" if card["is_aggregate"] else "")
        _fmt_cell(ws, r, 1, name)
        _fmt_cell(ws, r, 2, card["mid"])
        _fmt_cell(ws, r, 3, card["source"])
        _fmt_cell(ws, r, 4, card["total_spend"], NUM_FMT)
        _fmt_cell(ws, r, 5, card["total_pay"], NUM_FMT)
        _fmt_cell(ws, r, 6, card["total_roi"], "0.00")
        _fmt_cell(ws, r, 7, card["total_net_pay"], NUM_FMT)
        _fmt_cell(ws, r, 8, card["total_net_roi"], "0.00")
        r += 1
    _set_widths(ws, [14, 10, 14, 14, 14, 12, 12, 12, 11, 11, 11, 11, 12])
    ws.freeze_panes = "A4"


def build_materials(ws, weeks, mats, stats):
    cards = material_cards(weeks, mats, stats)
    _title(ws, "素材生命周期明细（共 %d 个素材，按累计消耗降序）" % len(cards), 32)
    headers = ["素材名称", "AI标识", "素材ID", "渠道", "产品", "作者", "混剪达人", "素材编号", "创作日期",
               "来源", "时长", "创建时间", "首次出现", "在投周数",
               "阶段", "预警", "累计消耗", "累计成交金额", "累计净成交", "累计订单数",
               "整体ROI", "净成交ROI", "最新周", "本周消耗", "本周成交金额", "本周净成交",
               "本周净成交ROI", "本周退款率", "本周结算率",
               "消耗环比", "成交金额环比", "整体ROI环比"]
    r = 2
    _header(ws, r, headers); r += 1
    for card in cards:
        name = card["name"] + ("（聚合）" if card["is_aggregate"] else "")
        ai_txt = card["ai_tag"] if card["is_ai"] else ("聚合" if card["is_aggregate"] else "")
        vals = [name, ai_txt, card["mid"], card["channel"], card["product"], card["author"],
                card["mix_author"], card["material_no"], card["created_date"],
                card["source"], card["duration"], card["created_at"],
                card["first_seen"], card["active_weeks"], card["phase"],
                "、".join(card["flags"]), card["total_spend"], card["total_pay"],
                card["total_net_pay"], card["total_order_count"], card["total_roi"],
                card["total_net_roi"], card["last_week"], card["last_spend"],
                card["last_pay"], card["last_net_pay"], card["last_net_roi"],
                card["last_refund_rate_1h"], card["last_settle_rate"],
                _pct(card["spend_delta"]), _pct(card["pay_delta"]), _pct(card["roi_delta"])]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            if j in (17, 18, 19, 24, 25, 26):
                c.number_format = NUM_FMT
            if j in (20,):
                c.number_format = INT_FMT
            if j in (21, 22, 27):
                c.number_format = "0.00"
            if j in (28, 29, 30, 31, 32):
                c.number_format = PCT_FMT
            if j == 15:
                c.alignment = CENTER
        r += 1
    widths = [46, 12, 20, 8, 12, 8, 14, 8, 12, 12, 8, 20, 12, 9, 10, 14, 13, 13, 13, 11, 9, 9, 12, 12, 12, 12, 12, 10, 10, 10, 12, 10]
    _set_widths(ws, widths)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:AF{r - 1}"


def _new_sheet(ws, report, title, week_end_map):
    _title(ws, title, 9)
    r = 3
    total_new = sum(len(w["items"]) for w in report)
    for idx, w in enumerate(report, start=1):
        items = w["items"]
        if not items:
            continue
        keep = [i for i in items if i["keep"]]
        keep_rate = len(keep) / len(items) if items else 0
        rois = [i["first_roi"] for i in items if i["first_roi"] is not None]
        avg_roi = sum(rois) / len(rois) if rois else None
        wk_end = week_end_map.get(w["week"], w["week"])
        head = f"第{idx}周（{w['week']} ~ {wk_end}） · 新增 {len(items)} 个 · 次周继续投放 {len(keep)} 个 ({keep_rate:.0%})"
        if avg_roi is not None:
            head += f" · 首周平均ROI {avg_roi:.2f}"
        ws.cell(row=r, column=1, value=head).font = SECTION_FONT
        r += 1
        _header(ws, r, ["素材名称", "素材ID", "创建时间", "首周消耗", "首周ROI", "次周继续", "次周消耗", "次周ROI"]); r += 1
        for i in items:
            _fmt_cell(ws, r, 1, i["name"])
            _fmt_cell(ws, r, 2, i["mid"])
            _fmt_cell(ws, r, 3, i["created_at"])
            _fmt_cell(ws, r, 4, i["first_spend"], NUM_FMT)
            _fmt_cell(ws, r, 5, i["first_roi"], "0.00")
            _fmt_cell(ws, r, 6, "是" if i["keep"] else ("否" if i["keep"] is False else "待观察"))
            _fmt_cell(ws, r, 7, i["next_spend"], NUM_FMT)
            _fmt_cell(ws, r, 8, i["next_roi"], "0.00")
            r += 1
        r += 1
    ws.cell(row=r, column=1, value=f"合计新增素材 {total_new} 个（按周去重后）").font = Font(name="微软雅黑", size=10, bold=True)
    _set_widths(ws, [46, 20, 20, 12, 10, 10, 12, 10])


def build_revisions(ws, account):
    conn = get_conn(account)
    rows = conn.execute(
        "SELECT rl.*, m.name FROM revision_log rl LEFT JOIN materials m ON m.material_key=rl.material_key"
        " ORDER BY rl.id").fetchall()
    conn.close()
    _title(ws, f"修订记录（官方回调/重复导入覆盖痕迹，共 {len(rows)} 条）", 7)
    r = 2
    _header(ws, r, ["时间", "素材名称", "天", "字段", "旧值", "新值"]); r += 1
    for row in rows:
        _fmt_cell(ws, r, 1, row["revised_at"])
        _fmt_cell(ws, r, 2, row["name"])
        _fmt_cell(ws, r, 3, row["period_start"])
        _fmt_cell(ws, r, 4, row["field"])
        _fmt_cell(ws, r, 5, row["old_value"])
        _fmt_cell(ws, r, 6, row["new_value"])
        r += 1
    _set_widths(ws, [20, 46, 12, 16, 18, 18])


def generate_excel_report(account=None):
    account = account or DEFAULT_ACCOUNT
    out_path = excel_out_path(account)
    os.makedirs(OUT_DIR, exist_ok=True)
    weeks, mats, stats = load_all(account)
    wb = Workbook()

    ws = wb.active
    ws.title = "汇总总览"
    build_overview(ws, weeks, mats, stats, account)

    ws = wb.create_sheet("素材生命周期明细")
    build_materials(ws, weeks, mats, stats)

    created_report, firstseen_report = new_week_report(weeks, mats, stats, {c["key"]: c for c in material_cards(weeks, mats, stats)})
    week_end_map = get_week_end_map(account)
    ws = wb.create_sheet("新增素材-创建口径")
    _new_sheet(ws, created_report, "每周新增素材（口径A：素材创建时间落在该周）", week_end_map)
    ws = wb.create_sheet("新增素材-首现口径")
    _new_sheet(ws, firstseen_report, "每周新增素材（口径B：首次出现在周报中）", week_end_map)

    ws = wb.create_sheet("修订记录")
    build_revisions(ws, account)

    wb.save(out_path)
    print(f"  Excel 周报已生成: {out_path}")


def generate_all_excel_reports():
    """为所有已入库的账户生成周报。"""
    for account in discover_accounts():
        if account_has_data(account):
            generate_excel_report(account)


if __name__ == "__main__":
    generate_all_excel_reports()
