# -*- coding: utf-8 -*-
"""验证：DB 计数、汇总数字、Excel 与看板输出完整性。

数据源为天级（daily_stats），周报表为 ISO 自然周聚合。
"""
import os
import sys
import json

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE, "monitor"))

from db import get_conn
from openpyxl import load_workbook

LINES = []
def P(*a):
    LINES.append(" ".join(str(x) for x in a))

conn = get_conn()
P("== 数据库计数 ==")
for t in ["materials", "daily_stats", "import_log", "revision_log"]:
    n = conn.execute(f"SELECT COUNT(*) c FROM {t}").fetchone()["c"]
    P(f"  {t}: {n}")
tot_spend = conn.execute("SELECT SUM(spend_total) s FROM daily_stats").fetchone()["s"]
tot_pay = conn.execute("SELECT SUM(pay_total) s FROM daily_stats").fetchone()["s"]
tot_net = conn.execute("SELECT SUM(net_pay) s FROM daily_stats").fetchone()["s"]
tot_ord = conn.execute("SELECT SUM(order_count) s FROM daily_stats").fetchone()["s"]
P(f"  总消耗: {tot_spend:,.2f}  总成交金额: {tot_pay:,.2f}  总净成交: {tot_net:,.2f}  总订单: {tot_ord:,.0f}")
P(f"  整体ROI: {tot_pay/tot_spend:.4f}  净成交ROI: {tot_net/tot_spend:.4f}")
days = [r["period_start"] for r in conn.execute("SELECT DISTINCT period_start FROM daily_stats ORDER BY period_start")]
P(f"  天数: {len(days)}  首个: {days[0]}  末个: {days[-1]}")
agg = conn.execute("SELECT COUNT(*) c FROM materials WHERE is_aggregate=1").fetchone()["c"]
P(f"  聚合行素材: {agg}")
conn.close()

P("")
P("== Excel 周报 ==")
xlsx = os.path.join(BASE, "output", "素材监控周报.xlsx")
wb = load_workbook(xlsx, read_only=True)
P(f"  Sheet: {wb.sheetnames}")
ws = wb["汇总总览"]
# 读取累计行
found = False
for row in ws.iter_rows(values_only=True):
    if row and row[0] == "累计":
        P(f"  累计行: {row}")
        found = True
wb.close()

P("")
P("== 网页看板 data.js ==")
djs = os.path.join(BASE, "output", "dashboard", "data.js")
with open(djs, encoding="utf-8") as fh:
    text = fh.read()
data = json.loads(text[len("window.DASH_DATA = "):-1])
P(f"  生成时间: {data['generated_at']}")
P(f"  周数: {len(data['weeks'])}  趋势行: {len(data['trend'])}  素材卡: {len(data['materials'])}")
P(f"  总量: 消耗{data['totals']['spend']:,.2f} 成交金额{data['totals']['pay']:,.2f} "
  f"净成交{data['totals']['net_pay']:,.2f} 订单{data['totals']['order_count']:,.0f} 素材数{data['totals']['material_count']}")
last = data["trend"][-1]
P(f"  末周({last['week']}): 消耗{last['spend']:,.2f} 成交金额{last['pay']:,.2f} ROI{last['roi']:.2f} 环比{last.get('spend_delta')}")
created_tot = sum(len(w["items"]) for w in data["new_created"])
first_tot = sum(len(w["items"]) for w in data["new_firstseen"])
P(f"  新增口径A(创建)合计: {created_tot}  口径B(首现)合计: {first_tot}")
rev_n = len(data["revisions"])
P(f"  修订记录: {rev_n} 条")
with open(os.path.join(BASE, "verify_out.txt"), "w", encoding="utf-8") as fh:
    fh.write("\n".join(LINES))
print("OK")
